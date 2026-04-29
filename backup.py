#!/usr/bin/env python3
"""
Bakery App Database Backup Script
备份 TiDB Cloud 数据库

Usage:
    python backup.py backup              # 完整备份 (SQL格式)
    python backup.py export              # 导出 Excel 格式
    python backup.py restore <file>      # 恢复备份
    python backup.py list                # 列出所有备份
    python backup.py tables              # 列出所有表
"""

import os
import sys
import json
import argparse
from datetime import datetime
from pathlib import Path

try:
    import mysql.connector
    from mysql.connector import Error
except ImportError:
    print("请安装 mysql-connector-python: pip install mysql-connector-python")
    sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    print("请安装 openpyxl: pip install openpyxl")
    sys.exit(1)


def load_env():
    """从 .env 文件加载环境变量"""
    env_path = Path(__file__).parent / ".env"
    if not env_path.exists():
        env_path = Path(__file__).parent / "backend" / ".env"

    env_vars = {}
    if env_path.exists():
        with open(env_path) as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#"):
                    key, value = line.split("=", 1)
                    env_vars[key] = value
    return env_vars


def get_connection(env):
    """创建数据库连接"""
    try:
        connection = mysql.connector.connect(
            host=env.get("DB_HOST"),
            port=int(env.get("DB_PORT", 4000)),
            user=env.get("DB_USER"),
            password=env.get("DB_PASSWORD"),
            database=env.get("DB_NAME", "bakery"),
        )
        return connection
    except Error as e:
        print(f"连接数据库失败: {e}")
        sys.exit(1)


def get_all_tables(connection):
    """获取所有表"""
    cursor = connection.cursor()
    cursor.execute("SHOW TABLES")
    tables = [row[0] for row in cursor.fetchall()]
    cursor.close()
    return tables


def backup_table(cursor, table_name):
    """备份单个表，返回 CREATE 和 INSERT 语句"""
    statements = []

    # CREATE TABLE
    cursor.execute(f"SHOW CREATE TABLE `{table_name}`")
    create_stmt = cursor.fetchone()[1]
    statements.append(f"\n\n-- Table: {table_name}\n")
    statements.append(f"DROP TABLE IF EXISTS `{table_name}`;\n")
    statements.append(create_stmt + ";\n")

    # INSERT data
    cursor.execute(f"SELECT * FROM `{table_name}`")
    rows = cursor.fetchall()
    if rows:
        cursor.execute(f"SHOW COLUMNS FROM `{table_name}`")
        columns = [col[0] for col in cursor.fetchall()]

        for row in rows:
            values = []
            for val in row:
                if val is None:
                    values.append("NULL")
                elif isinstance(val, (int, float)):
                    values.append(str(val))
                elif isinstance(val, datetime):
                    values.append(f"'{val.strftime('%Y-%m-%d %H:%M:%S')}'")
                elif isinstance(val, bytes):
                    values.append(f"0x{val.hex()}")
                else:
                    escaped = str(val).replace("\\", "\\\\").replace("'", "\\'")
                    values.append(f"'{escaped}'")
            statements.append(
                f"INSERT INTO `{table_name}` ({', '.join(columns)}) VALUES ({', '.join(values)});\n"
            )

    return "".join(statements)


def backup_database(backup_dir=None):
    """备份整个数据库"""
    env = load_env()

    if backup_dir is None:
        backup_dir = Path(__file__).parent / "backups"
    else:
        backup_dir = Path(backup_dir)

    backup_dir.mkdir(exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_file = backup_dir / f"bakery_backup_{timestamp}.sql"

    print(f"正在备份数据库: {env.get('DB_NAME')}")
    print(f"备份文件: {backup_file}")

    connection = get_connection(env)
    cursor = connection.cursor()

    tables = get_all_tables(connection)
    print(f"找到 {len(tables)} 个表: {', '.join(tables)}")

    with open(backup_file, "w", encoding="utf-8") as f:
        f.write(f"-- Bakery App Database Backup\n")
        f.write(f"-- Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"-- Database: {env.get('DB_NAME')}\n")
        f.write(f"-- Tables: {', '.join(tables)}\n\n")

        for i, table in enumerate(tables):
            print(f"  备份中 [{i + 1}/{len(tables)}]: {table}")
            f.write(backup_table(cursor, table))

    cursor.close()
    connection.close()

    # 同时保存备份元数据
    meta_file = backup_dir / f"bakery_backup_{timestamp}_meta.json"
    with open(meta_file, "w") as f:
        json.dump(
            {
                "timestamp": timestamp,
                "date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "database": env.get("DB_NAME"),
                "tables": tables,
                "backup_file": str(backup_file),
            },
            f,
            indent=2,
        )

    print(f"\n备份完成! 文件: {backup_file}")
    return str(backup_file)


def export_excel(backup_dir=None):
    """导出数据库为 Excel 格式"""
    env = load_env()

    if backup_dir is None:
        backup_dir = Path(__file__).parent / "backups"
    else:
        backup_dir = Path(backup_dir)

    backup_dir.mkdir(exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_file = backup_dir / f"bakery_export_{timestamp}.xlsx"

    print(f"正在导出数据库: {env.get('DB_NAME')}")
    print(f"导出文件: {excel_file}")

    connection = get_connection(env)
    cursor = connection.cursor()

    tables = get_all_tables(connection)
    print(f"找到 {len(tables)} 个表: {', '.join(tables)}")

    wb = Workbook()
    wb.remove(wb.active)  # 移除默认 sheet

    # 样式定义
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="8B5A2B", end_color="8B5A2B", fill_type="solid"
    )
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for i, table in enumerate(tables):
        print(f"  导出中 [{i + 1}/{len(tables)}]: {table}")

        # 获取表数据
        cursor.execute(f"SELECT * FROM `{table}`")
        rows = cursor.fetchall()
        columns = [col[0] for col in cursor.description] if cursor.description else []

        if not columns:
            continue

        ws = wb.create_sheet(title=table[:31])  # Excel sheet name max 31 chars

        # 写入表头
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 写入数据
        for row_idx, row in enumerate(rows, 2):
            for col_idx, value in enumerate(row, 1):
                if isinstance(value, datetime):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.number_format = "YYYY-MM-DD HH:MM:SS"
                elif value is None:
                    cell = ws.cell(row=row_idx, column=col_idx, value="")
                else:
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

        # 自动调整列宽
        for col_idx, col_name in enumerate(columns, 1):
            max_length = len(str(col_name))
            for row in rows:
                if row[col_idx - 1] is not None:
                    max_length = max(max_length, len(str(row[col_idx - 1])))
            ws.column_dimensions[
                chr(64 + col_idx) if col_idx <= 26 else "A" + chr(64 + col_idx - 26)
            ].width = min(max_length + 2, 50)

    cursor.close()
    connection.close()

    wb.save(excel_file)
    print(f"\n导出完成! 文件: {excel_file}")
    return str(excel_file)


def restore_database(backup_file):
    """从备份文件恢复数据库"""
    env = load_env()
    backup_path = Path(backup_file)

    if not backup_path.exists():
        print(f"备份文件不存在: {backup_file}")
        sys.exit(1)

    print(f"正在恢复数据库: {env.get('DB_NAME')}")
    print(f"备份文件: {backup_path}")

    connection = get_connection(env)
    cursor = connection.cursor()

    with open(backup_path, "r", encoding="utf-8") as f:
        sql_content = f.read()

    # 分割并执行每个语句
    statements = []
    current_stmt = []
    in_string = False
    string_char = None

    for char in sql_content:
        if not in_string:
            if char in ("'", '"'):
                in_string = True
                string_char = char
            elif char == ";":
                if current_stmt:
                    statements.append("".join(current_stmt))
                    current_stmt = []
            else:
                current_stmt.append(char)
        else:
            if char == string_char:
                in_string = False
                string_char = None
            current_stmt.append(char)

    total = len(statements)
    print(f"共 {total} 条 SQL 语句")

    for i, stmt in enumerate(statements):
        stmt = stmt.strip()
        if stmt and not stmt.startswith("--") and not stmt.startswith("#"):
            try:
                cursor.execute(stmt)
                if (i + 1) % 50 == 0:
                    print(f"  执行中 [{i + 1}/{total}]")
            except Error as e:
                print(f"  执行失败 [{i + 1}]: {e}")
                print(f"  语句: {stmt[:100]}...")

    connection.commit()
    cursor.close()
    connection.close()

    print("\n恢复完成!")


def list_backups(backup_dir=None):
    """列出所有备份"""
    if backup_dir is None:
        backup_dir = Path(__file__).parent / "backups"
    else:
        backup_dir = Path(backup_dir)

    if not backup_dir.exists():
        print("没有找到备份目录")
        return

    backup_files = list(backup_dir.glob("bakery_backup_*.sql"))

    if not backup_files:
        print("没有找到备份文件")
        return

    print(f"找到 {len(backup_files)} 个备份:\n")
    print(f"{'日期':<20} {'时间':<10} {'大小':<12} {'文件'}")
    print("-" * 80)

    for bf in sorted(backup_files, reverse=True):
        stat = bf.stat()
        dt = datetime.fromtimestamp(stat.st_mtime)
        size = stat.st_size
        if size > 1024 * 1024:
            size_str = f"{size / 1024 / 1024:.1f} MB"
        elif size > 1024:
            size_str = f"{size / 1024:.1f} KB"
        else:
            size_str = f"{size} B"
        print(
            f"{dt.strftime('%Y-%m-%d'):<20} {dt.strftime('%H:%M:%S'):<10} {size_str:<12} {bf.name}"
        )


def list_tables():
    """列出所有表"""
    env = load_env()
    connection = get_connection(env)
    cursor = connection.cursor()

    tables = get_all_tables(connection)

    print(f"数据库: {env.get('DB_NAME')}")
    print(f"共 {len(tables)} 个表:\n")

    for table in tables:
        cursor.execute(f"SELECT COUNT(*) FROM `{table}`")
        count = cursor.fetchone()[0]
        print(f"  {table:<40} {count:>5} 行")

    cursor.close()
    connection.close()


def main():
    parser = argparse.ArgumentParser(description="Bakery App 数据库备份工具")
    subparsers = parser.add_subparsers(dest="command", help="子命令")

    subparsers.add_parser("backup", help="备份数据库 (SQL格式)")
    subparsers.add_parser("export", help="导出 Excel 格式")
    subparsers.add_parser("list", help="列出所有备份")
    subparsers.add_parser("tables", help="列出所有表")

    restore_parser = subparsers.add_parser("restore", help="恢复备份")
    restore_parser.add_argument("file", help="备份文件路径")

    args = parser.parse_args()

    if args.command == "backup":
        backup_database()
    elif args.command == "export":
        export_excel()
    elif args.command == "restore":
        restore_database(args.file)
    elif args.command == "list":
        list_backups()
    elif args.command == "tables":
        list_tables()
    else:
        parser.print_help()


if __name__ == "__main__":
    main()
